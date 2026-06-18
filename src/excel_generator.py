import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from src.config import DATA_PROCESSED_DIR, OUTPUTS_DIR

def sip_value(monthly_amount, annual_cagr, years=5):
    """Calculate the final value of a monthly SIP using future value of annuity due formula."""
    if annual_cagr is None or pd.isna(annual_cagr):
        return 0, 0
    n = int(years * 12)  # total months
    r = (annual_cagr / 100.0) / 12.0  # monthly rate
    
    if r == 0:
        fv = monthly_amount * n
    else:
        # Future value of annuity due: paid at beginning of period
        fv = monthly_amount * (((1 + r)**n - 1) / r) * (1 + r)
        
    total_invested = monthly_amount * n
    gain = fv - total_invested
    return round(fv), round(gain)

def style_worksheet(ws, header_title=None):
    """Apply premium styling: fonts, borders, grid lines, and alignments to a worksheet."""
    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True
    
    # Premium Font Family
    font_family = "Segoe UI"
    
    # Curated Styling Palettes
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")  # Navy Blue
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10, bold=False, color="333333")
    
    thin_border_side = Side(style='thin', color="E0E0E0")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Style Header Row (Row 1)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[1].height = 28
    
    # Style Data Rows
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 20
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = thin_border
            
            # Smart alignment and formatting based on data type / values
            val = cell.value
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                # Format numbers based on column headers
                header_text = str(ws.cell(row=1, column=col).value).lower()
                if "cagr" in header_text or "alpha" in header_text or "volatility" in header_text:
                    cell.number_format = '0.00"%"'
                elif "sharpe" in header_text:
                    cell.number_format = '0.00'
                elif "score" in header_text:
                    cell.number_format = '0.0'
                elif "value" in header_text or "gain" in header_text or "invested" in header_text:
                    cell.number_format = '[$Rs-439] #,##0'  # Indian Rupee Format
            elif str(val).startswith("S -") or str(val).startswith("A -") or str(val).startswith("B -") or str(val).startswith("C -") or str(val).startswith("D -"):
                cell.alignment = Alignment(horizontal="center", vertical="center")
                # Apply soft colors to grades for readability
                grade = str(val)
                if "S -" in grade:
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Soft Green
                elif "A -" in grade:
                    cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # Soft Blue
                elif "B -" in grade:
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Soft Yellow
                elif "C -" in grade:
                    cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # Soft Orange
                elif "D -" in grade:
                    cell.fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")  # Soft Red
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    # Auto-fit Column Widths with padding
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                # Add extra padding for percentage sign and rupee formats
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

def run_excel_generation():
    """Generates the multi-sheet Excel workbook outputs/MF_Report.xlsx."""
    print("\nGenerating formatted Excel report workbook...")
    
    scored_path = os.path.join(DATA_PROCESSED_DIR, "fund_scores_final.csv")
    if not os.path.exists(scored_path):
        print("[ERROR] Required CSV file fund_scores_final.csv missing! Run statistical_analyser first.")
        return
        
    df = pd.read_csv(scored_path)
    
    wb = Workbook()
    
    # --- SHEET 1: Fund Rankings ---
    ws1 = wb.active
    ws1.title = "Fund Rankings"
    df_rankings = df.sort_values("Score", ascending=False).copy()
    columns_rankings = [
        "Fund_Name", "Category", "Score", "Score_Grade", "Rank_in_Category",
        "CAGR_5yr", "CAGR_3yr", "CAGR_1yr", "Sharpe_Ratio", "Volatility", "Alpha_5yr"
    ]
    # Rename for Excel display
    df_rankings_sheet = df_rankings[columns_rankings].copy()
    df_rankings_sheet.columns = [
        "Fund Name", "Category", "Composite Score", "Grade", "Rank in Category",
        "5-Year CAGR", "3-Year CAGR", "1-Year CAGR", "Sharpe Ratio", "Volatility", "5-Year Alpha"
    ]
    
    # Write to sheet
    for r in dataframe_to_rows(df_rankings_sheet):
        ws1.append(r)
    style_worksheet(ws1)
    
    # --- SHEET 2: Category Analysis ---
    ws2 = wb.create_sheet(title="Category Analysis")
    df_cat = df.groupby("Category")[["CAGR_5yr", "CAGR_3yr", "CAGR_1yr", "Sharpe_Ratio", "Volatility", "Alpha_5yr"]].mean().reset_index()
    df_cat.columns = [
        "Category", "Avg 5-Year CAGR", "Avg 3-Year CAGR", "Avg 1-Year CAGR", 
        "Avg Sharpe Ratio", "Avg Volatility", "Avg 5-Year Alpha"
    ]
    for r in dataframe_to_rows(df_cat):
        ws2.append(r)
    style_worksheet(ws2)
    
    # --- SHEET 3: Alpha Report ---
    ws3 = wb.create_sheet(title="Alpha Report")
    df_alpha = df.sort_values("Alpha_5yr", ascending=False).copy()
    df_alpha_sheet = df_alpha[[
        "Fund_Name", "Category", "Alpha_5yr", "CAGR_5yr", "Bench_5yr", "Benchmark"
    ]].copy()
    df_alpha_sheet.columns = [
        "Fund Name", "Category", "5-Year Alpha", "5-Year Fund CAGR", "5-Year Benchmark CAGR", "Benchmark Ticker"
    ]
    for r in dataframe_to_rows(df_alpha_sheet):
        ws3.append(r)
    style_worksheet(ws3)
    
    # --- SHEET 4: Top Picks ---
    ws4 = wb.create_sheet(title="Top Picks")
    df_picks = df[df["Score_Grade"].isin(["S - Excellent", "A - Good"])].sort_values("Score", ascending=False).copy()
    df_picks_sheet = df_picks[columns_rankings].copy()
    df_picks_sheet.columns = [
        "Fund Name", "Category", "Composite Score", "Grade", "Rank in Category",
        "5-Year CAGR", "3-Year CAGR", "1-Year CAGR", "Sharpe Ratio", "Volatility", "5-Year Alpha"
    ]
    for r in dataframe_to_rows(df_picks_sheet):
        ws4.append(r)
    style_worksheet(ws4)
    
    # --- SHEET 5: SIP Calculator ---
    ws5 = wb.create_sheet(title="SIP Calculator")
    top10_funds = df.nlargest(10, "Score").copy()
    
    sip_amounts = [1000, 2000, 5000, 10000]
    sip_rows = []
    
    for _, fund in top10_funds.iterrows():
        row = {
            "Fund Name": fund["Fund_Name"],
            "Category": fund["Category"],
            "5-Year CAGR": fund["CAGR_5yr"],
            "Composite Score": fund["Score"]
        }
        for amt in sip_amounts:
            val, gain = sip_value(amt, fund["CAGR_5yr"], 5)
            row[f"Value (Rs {amt:,}/mo)"] = val
            row[f"Gain (Rs {amt:,}/mo)"] = gain
        sip_rows.append(row)
        
    df_sip = pd.DataFrame(sip_rows)
    for r in dataframe_to_rows(df_sip):
        ws5.append(r)
    style_worksheet(ws5)
    
    # --- SHEET 6: Raw Data ---
    ws6 = wb.create_sheet(title="Raw Data")
    df_raw = df.copy()
    # Write entire DataFrame (rename all columns nicely)
    for r in dataframe_to_rows(df_raw):
        ws6.append(r)
    style_worksheet(ws6)
    
    # Save Workbook
    out_xlsx_path = os.path.join(OUTPUTS_DIR, "MF_Report.xlsx")
    wb.save(out_xlsx_path)
    print(f"[SUCCESS] Formatted multi-sheet Excel report saved to: {out_xlsx_path}")

def dataframe_to_rows(df):
    """Helper to convert pandas DataFrame to rows suitable for appending, including headers."""
    yield list(df.columns)
    for row in df.itertuples(index=False):
        yield list(row)

if __name__ == "__main__":
    run_excel_generation()
